"""Seed rows for the data-entry workbook.

These are starting points for the curator, not field data. Everything here is drawn
from publicly documented patterns of hate speech against Iraqi minorities — the
Sinjar genocide and its aftermath, the Nineveh Plains displacement, the recurring
loyalty and authenticity smears — and every row is marked `seed` in its notes column
so it is distinguishable from anything the Duhok survey actually attested. A curator
should confirm, correct or delete each one against the real data before it is trusted.

Two decisions worth stating, because they shape what is here:

**Weighted towards what a keyword list cannot do.** The survey's own reference sheet
records that mockery is the most-reported form and "rarely uses a slur, so a
keyword-only system misses it almost entirely". A lexicon of slurs would score well on
the cases that were never the problem. So the tropes carry most of the weight, and
several lexicon entries exist mainly to hold a `never_flag_when` rule.

**Self-reference terms are not in the lexicon.** `ايزيدي`, `صابئة` and `شنكال` are
how communities name themselves and where they live, and the first draft of this file
put them here with severity 0 so a keyword pass could not flag them. The importer
refused — severity must be 1-10 — and it was right to, because the lexicon is a list
of terms that carry harm and there is no honest weight for a word that carries none.
They belong in the target-group pack's `self_reference_terms`, which is read by
src/packs/loader.py and exists for exactly this. Left in the lexicon they would have
needed a fake severity, and the next person to sort the sheet by weight would have
found the community's own name sitting in a list of slurs.

**Benign rows are not padding.** The importer warns if hateful examples outnumber
benign ones, and it is right to. A classifier trained or evaluated only on hate learns
that everything mentioning Yazidis is suspicious, which in a human-rights record means
false accusations against the people the system exists to protect. Several pairs here
are the *same phrase* under different parents, benign in one and hateful in the other.

Add:    python tools/seed_lexicon_rows.py docs/lexicon_data_entry_template.xlsx
Remove: python tools/seed_lexicon_rows.py docs/lexicon_data_entry_template.xlsx --clear
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

# Every row this script writes carries this prefix in its notes — or, on EXAMPLES
# which has no notes column, in `why`. It is what --clear finds, and what tells a
# curator which rows are a reconstruction rather than something the survey attested.
MARKER = "seed —"
SEED = f"{MARKER} verify against Duhok field data before trusting"

# term, target_groups, language, category, severity, is_explicit, variants,
# never_flag_when, is_regex, notes, added_by
LEXICON: list[tuple] = [
    # ── explicit, low ambiguity ───────────────────────────────────────────────
    (
        "كفار", "yazidi, christian-iraqi, sabian-mandaean, kakai", "ar",
        "dehumanization", 7, "yes", "كافر, كفره, الكفار",
        "academic, news_quotation, counter_speech", "no",
        f"{SEED} — religious exclusion used to place a community outside protection. "
        "Genuinely appears in theological discussion, so the never_flag rules matter.",
        "ankedo-agent",
    ),
    (
        "مشركين", "yazidi, sabian-mandaean", "ar", "dehumanization", 7, "yes",
        "مشرك, المشركين", "academic, news_quotation, counter_speech", "no",
        f"{SEED} — same mechanism as كفار.", "ankedo-agent",
    ),
    (
        "نجس", "yazidi, sabian-mandaean, christian-iraqi", "ar", "dehumanization", 8,
        "yes", "انجاس, نجسين, النجس", "academic, news_quotation, counter_speech", "no",
        f"{SEED} — ritual-impurity framing. Precedes exclusion demands: barring from "
        "wells, shops, food handling.",
        "ankedo-agent",
    ),
    (
        "اقتلوهم", "yazidi, christian-iraqi, shabak, sabian-mandaean", "ar",
        "incitement", 10, "yes", "اقتلوهن, يجب قتلهم, اقتلهم",
        "news_quotation, academic", "no",
        f"{SEED} — direct incitement; the ceiling of the scale.", "ankedo-agent",
    ),
    (
        "يجب ترحيلهم", "christian-iraqi, yazidi, shabak, turkmen-iraqi", "ar",
        "exclusion_call", 8, "yes", "ترحيلهم, يطلعون من العراق, لا مكان لهم بالعراق",
        "news_quotation, academic", "no",
        f"{SEED} — expulsion demand. Distinct from a slur: it names an action.",
        "ankedo-agent",
    ),

    # ── the loyalty smear, from the survey's own disloyalty term list ─────────
    (
        "عملاء", "christian-iraqi, yazidi, kurdish", "ar", "disloyalty", 6, "yes",
        "عميل, العملاء, عملاء الغرب, عملاء اسرائيل",
        "news_quotation, academic, counter_speech", "no",
        f"{SEED} — REFERENCE lists disloyalty terms as frequent in the survey. "
        "Treats a whole community as a foreign asset.",
        "ankedo-agent",
    ),
    (
        "الخونة", "christian-iraqi, yazidi, kurdish, faili-kurd", "ar", "disloyalty",
        6, "yes", "خونة, خائن, خيانة", "news_quotation, academic, counter_speech", "no",
        f"{SEED} — from the survey's term list. Very common in ordinary political "
        "argument, so it needs a target group present before it means anything.",
        "ankedo-agent",
    ),
    (
        "اتباع الانكليز", "christian-iraqi", "ar", "disloyalty", 6, "yes",
        "اتباع الإنكليز, صنيعة الانكليز, جايبينهم الانكليز",
        "news_quotation, academic", "no",
        f"{SEED} — from the survey's term list. Ties Assyrians to the British mandate "
        "to cast them as imported.",
        "ankedo-agent",
    ),
    (
        "صليبيين", "christian-iraqi", "ar", "foreignness", 7, "yes",
        "صليبي, الصليبيين", "news_quotation, academic, counter_speech", "no",
        f"{SEED} — recasts an indigenous community as a foreign invasion.",
        "ankedo-agent",
    ),
    (
        "دخلاء", "christian-iraqi, yazidi, shabak, kakai, faili-kurd", "ar",
        "foreignness", 6, "yes", "دخيل, دخلاء على العراق, مو اصليين",
        "news_quotation, academic", "no",
        f"{SEED} — already a surface form of the authenticity trope; kept here so the "
        "bare term is caught when the trope does not fire.",
        "ankedo-agent",
    ),

    # ── Kurdish-script terms ──────────────────────────────────────────────────
    (
        "بێ دین", "yazidi, christian-iraqi", "ku", "dehumanization", 7, "yes",
        "بێدین, بێ ئایین", "academic, news_quotation, counter_speech", "no",
        f"{SEED} — 'without religion'. Kurmanji/Sorani equivalent of the كفار framing.",
        "ankedo-agent",
    ),
    (
        "خوفرۆش", "kurdish, yazidi, christian-iraqi", "ku", "disloyalty", 6, "yes",
        "خۆفرۆش, خائن", "news_quotation, academic, counter_speech", "no",
        f"{SEED} — 'self-seller', the Kurdish traitor smear.", "ankedo-agent",
    ),

]

# name, description, target_groups, surface_forms, activation_topics,
# example (HATE), negative_example (BENIGN), counter_speech_example,
# severity, is_visual, requires_target_group, negation_cancels, notes
TROPES: list[tuple] = [
    (
        "Foreign agents / loyalty smear",
        "Casts a whole community as serving a foreign power — Israel, the West, the "
        "British mandate — so that its members can be treated as enemies rather than "
        "citizens. The words are ordinary political insults; what makes it hate speech "
        "is that the accusation is collective and inherited.",
        "christian-iraqi, yazidi, kurdish, faili-kurd",
        "عملاء, الخونة, اتباع الانكليز, عملاء اسرائيل, صنيعة",
        "اشوري, مسيحي, ايزيدي, كلداني, نينوى, شنكال",
        "المسيحية كلهم عملاء لاسرائيل ولازم ينحاسبون",
        "هاي الحكومة كلها عملاء",  # same insult, no group named
        "اتهام طائفة كاملة بالخيانة بسبب دينها هو خطاب كراهية، وليس نقداً سياسياً",
        6, "no", "yes", "yes",
        f"{SEED} — requires_target_group is essential: 'الخونة' is everyday political "
        "argument in Iraq and flagging it unconditionally would bury the queue.",
    ),
    (
        "Ritual impurity",
        "Frames the community as physically or spiritually contaminating. Historically "
        "the step before exclusion demands — from wells, shops, food handling, schools "
        "— so it should be read as a precursor to exclusion rather than a mere insult.",
        "yazidi, sabian-mandaean, christian-iraqi, kakai",
        "نجس, انجاس, ما ناكل من ايدهم, ما نشرب وياهم",
        "ايزيدي, صابئة, مندائي, مسيحي, اشوري",
        "لا تشترون منهم، هذوله نجس ما يصير تاكل من ايدهم",
        "الماي هنا نجس ما ينشرب",  # impurity of a thing, not a people
        "وصف الناس بالنجاسة بسبب دينهم هو تمهيد للإقصاء، مو رأي",
        8, "no", "yes", "yes",
        f"{SEED} — the benign example matters: نجس is an ordinary word for unclean "
        "and only becomes hate speech when a community is its object.",
    ),
    (
        "Mockery of religious practice",
        "Ridicules a minority's rites, dress, pilgrimage or holy sites. Rarely contains "
        "a slur, which is why a keyword system misses it — the REFERENCE sheet records "
        "mockery as the most-reported form in the survey. The insult is carried by "
        "framing and by what it is a reply to.",
        "yazidi, christian-iraqi, sabian-mandaean, kakai",
        "",  # deliberately none — this trope is not lexical
        "لالش, شنكال, معبد, طقوس, عماد, تعميد, صليب, ايزيدي, مسيحي, صابئة",
        "شوفوا هاي الطقوس المضحكة، يدورون حول حجر ويسمونها عبادة",
        "زرت لالش هذا الاسبوع، مكان جميل وهادئ",
        "السخرية من طقوس الناس مو نكتة، هاي إهانة لجماعة كاملة",
        5, "no", "yes", "no",
        f"{SEED} — no surface_forms on purpose. If this trope is written as a keyword "
        "list it will match the ordinary vocabulary of religious life and flag the "
        "community's own posts about itself.",
    ),
    (
        "Genocide denial or minimisation",
        "Denies, downplays or justifies the 2014 Sinjar genocide, or recasts the "
        "victims as complicit. Distinct from historical debate by its function: it "
        "removes the grounds on which survivors ask for protection.",
        "yazidi",
        "مبالغة, ما صار شي, يستاهلون, كذبة",
        "شنكال, سنجار, داعش, ٢٠١٤, سبايا, ابادة",
        "قصة الإبادة مبالغة سياسية، ما صار بشنكال ولا شي",
        "المؤرخون يناقشون أعداد الضحايا في شنكال",
        "إنكار ما صار بشنكال إهانة للناجيات وللي ماتوا",
        9, "no", "yes", "yes",
        f"{SEED} — the benign example is deliberately close: discussion of casualty "
        "figures is legitimate scholarship and must not be flagged.",
    ),
    (
        "Demographic replacement claim",
        "Alleges that a minority is seizing land or altering the demographic balance "
        "of an area, usually the Nineveh Plains or Sinjar. Supplies the justification "
        "for expulsion demands, so it often appears just before them.",
        "christian-iraqi, shabak, yazidi, kurdish",
        "يستولون, تغيير ديموغرافي, يشترون الاراضي",
        "نينوى, سهل نينوى, برطلة, شنكال, الموصل",
        "هذوله يشترون اراضينا ويغيرون ديموغرافية المنطقة، لازم نوقفهم",
        "تقرير عن التغير السكاني في سهل نينوى بعد ٢٠١٤",
        "الناس ترجع لبيوتها مو استيلاء، هذي ارضهم اصلا",
        7, "no", "yes", "yes",
        f"{SEED} — pairs with exclusion_call; when both fire on one thread the thread "
        "is worth expanding.",
    ),
    (
        "Denial of indigenous status",
        "Asserts the community arrived recently, was planted by a foreign power, or is "
        "not a real people — stripping the standing from which it can make claims. "
        "Overlaps the existing authenticity trope but reaches the Shabak, Kaka'i and "
        "Faili Kurds, whom that one does not name.",
        "shabak, kakai, faili-kurd, sabian-mandaean, turkmen-iraqi",
        "مو شعب, ما الهم اصل, جايين من ايران, مو عراقيين",
        "شبك, كاكائي, فيلي, صابئة, تركمان, برطلة",
        "الشبك مو عراقيين اصلا، جايين من ايران ولا الهم حق بالمنطقة",
        "الشبك من مكونات العراق وليهم تاريخ طويل بسهل نينوى",
        "كل مكونات العراق اصيلة، والتشكيك باصل الناس خطاب كراهية",
        7, "no", "yes", "yes",
        f"{SEED} — added because the existing authenticity trope covers only "
        "christian-iraqi, leaving four groups with no trope at all.",
    ),
    (
        "Mocking imagery of dress or symbols",
        "An image whose payload is ridicule of religious dress, a symbol or a holy "
        "site — a defaced sun symbol, a mocking edit of clerical dress, a caption "
        "photographed onto a shrine. The caption is often innocuous by design, because "
        "the poster knows a text classifier reads only the caption.",
        "yazidi, christian-iraqi, sabian-mandaean, kakai",
        "",
        "لالش, صليب, معبد, طاووس ملك, رجل دين, شماس",
        "IMAGE: a defaced Yazidi sun symbol with a mocking caption",
        "IMAGE: a photograph of Lalish published by a news outlet",
        "IMAGE: a post reproducing the mocking image in order to condemn it",
        7, "yes", "yes", "no",
        f"{SEED} — is_visual: the verdict has to come from the image, and a text-only "
        "pass on the caption will clear it.",
    ),
]

# post_url, parent_post_text, row_type, comment_text, target_group, dialect,
# label, category, severity, why, demonstrates_trope, annotator_1, annotator_2,
# hard_case
_U = "https://facebook.com/seed/posts/"
EXAMPLES: list[tuple] = [
    # ── pair 1: the loyalty smear needs a target present ──────────────────────
    (
        f"{_U}101", "خبر: افتتاح كنيسة بعد ترميمها في سهل نينوى", "post",
        "خبر: افتتاح كنيسة بعد ترميمها في سهل نينوى", "christian-iraqi", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "A news post about a church reopening. Nothing here is an attack; it is the "
        "parent that gives the comments below their meaning.",
        "", "benign", "benign", "no",
    ),
    (
        f"{_U}101", "خبر: افتتاح كنيسة بعد ترميمها في سهل نينوى", "comment",
        "هذوله عملاء الغرب، شنو يريدون بالعراق", "christian-iraqi", "iraqi",
        "hate", "disloyalty", 6,
        f"{MARKER} " "Collective accusation of foreign allegiance, attached to a community named by "
        "the parent post. The insult is inherited rather than earned — no act is "
        "alleged, only membership.",
        "Foreign agents / loyalty smear", "hate", "hate", "no",
    ),
    (
        f"{_U}102", "منشور سياسي عن أداء الحكومة", "comment",
        "هذوله عملاء وما يفكرون بالبلد", "", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "HARD CASE — nearly the same words as the row above, under a political post "
        "with no community named. Ordinary Iraqi political abuse. The trope requires a "
        "target group precisely so this is not flagged.",
        "", "benign", "benign", "yes",
    ),

    # ── pair 2: impurity, and the same word about a thing ─────────────────────
    (
        f"{_U}103", "صور من احتفال ديني للصابئة المندائيين على النهر", "post",
        "صور من احتفال ديني للصابئة المندائيين على النهر", "sabian-mandaean", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "A community photographing its own ceremony.", "", "benign", "benign", "no",
    ),
    (
        f"{_U}103", "صور من احتفال ديني للصابئة المندائيين على النهر", "comment",
        "لا تشربون من هذا النهر بعد اليوم، هذوله نجس", "sabian-mandaean", "iraqi",
        "hate", "dehumanization", 8,
        f"{MARKER} " "Ritual-impurity framing aimed at people, followed immediately by an avoidance "
        "instruction. This is the shape that precedes exclusion from shared resources.",
        f"{MARKER} " "Ritual impurity", "hate", "hate", "no",
    ),
    (
        f"{_U}104", "شكوى من تلوث مياه الشرب في الحي", "comment",
        "الماي وصلنا نجس، ما ينشرب أبدا", "", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "HARD CASE — نجس in its ordinary sense, about water. No community is its "
        "object. A keyword list flags this; the trope does not.",
        "", "benign", "benign", "yes",
    ),

    # ── pair 3: mockery, the form with no slur in it ──────────────────────────
    (
        f"{_U}105", "مراسم إيزيدية في لالش بمناسبة العيد", "post",
        "مراسم إيزيدية في لالش بمناسبة العيد", "yazidi", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "Neutral description of a religious ceremony.", "", "benign", "benign", "no",
    ),
    (
        f"{_U}105", "مراسم إيزيدية في لالش بمناسبة العيد", "comment",
        "شوفوا هاي الطقوس المضحكة، يدورون حول حجر ويسمونها عبادة", "yazidi", "iraqi",
        "hate", "mockery", 5,
        f"{MARKER} " "Contains no slur at all — a keyword system clears it completely. The abuse is "
        "in ridiculing the practice as such, and the target is fixed by the parent.",
        "Mockery of religious practice", "hate", "hate", "yes",
    ),
    (
        f"{_U}105", "مراسم إيزيدية في لالش بمناسبة العيد", "comment",
        "زرت لالش قبل سنة، مكان هادئ وناسه طيبين", "yazidi", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "Same parent, same vocabulary of religious life, warm intent. Included so the "
        "activation topics cannot be treated as sufficient on their own.",
        "", "benign", "benign", "no",
    ),

    # ── pair 4: counter-speech quoting the libel ──────────────────────────────
    (
        f"{_U}106", "منشور عن الإيزيديين وحقوقهم", "comment",
        "يسمونهم عبدة الشيطان وهذا افتراء، الإيزيدية دين قديم ومحترم", "yazidi", "iraqi",
        "benign", "counter_speech", 0,
        f"{MARKER} " "HARD CASE — carries the slur عبدة الشيطان verbatim, and is a rejection of it. "
        "The never_flag_when rule `counter_speech` exists for exactly this row. "
        "Flagging it would put a defender in the evidence file.",
        "Devil-worship libel", "benign", "benign", "yes",
    ),

    # ── pair 5: genocide denial vs scholarship ────────────────────────────────
    (
        f"{_U}107", "ذكرى الإبادة في شنكال ٢٠١٤", "comment",
        "قصة الإبادة مبالغة سياسية، ما صار بشنكال ولا شي", "yazidi", "iraqi",
        "hate", "delegitimization", 9,
        f"{MARKER} " "Denial aimed at survivors on a commemoration post. Its function is to remove "
        "the grounds for protection, which is why it scores above ordinary mockery.",
        "Genocide denial or minimisation", "hate", "hate", "no",
    ),
    (
        f"{_U}107", "ذكرى الإبادة في شنكال ٢٠١٤", "comment",
        "الباحثون مختلفون على أعداد الضحايا، وهذا نقاش أكاديمي مستمر", "yazidi", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "HARD CASE — disputes figures on the same post without denying the event. "
        "Legitimate scholarship. The line is denial of the event, not debate about "
        "its measurement.",
        "", "benign", "benign", "yes",
    ),

    # ── pair 6: a group with no existing trope coverage ──────────────────────
    (
        f"{_U}108", "خبر عن مشاريع خدمية في برطلة", "comment",
        "الشبك مو عراقيين اصلا، جايين من ايران ولا الهم حق بالمنطقة", "shabak", "iraqi",
        "hate", "authenticity_denial", 7,
        f"{MARKER} " "Denies indigenous status in order to deny standing. Recorded because the "
        "Shabak had no trope covering them at all before this row.",
        f"{MARKER} " "Denial of indigenous status", "hate", "hate", "no",
    ),
    (
        f"{_U}108", "خبر عن مشاريع خدمية في برطلة", "comment",
        "برطلة مدينة جميلة واهلها من اطيب الناس", "shabak", "iraqi",
        "benign", "", 0,
        f"{MARKER} " "Ordinary praise under the same post.", "", "benign", "benign", "no",
    ),
]


def _sheet(wb, prefix: str):
    name = next((n for n in wb.sheetnames if n.upper().startswith(prefix)), None)
    if name is None:
        raise SystemExit(f"no sheet starting with {prefix!r} — is this the right workbook?")
    return wb[name]


def _first_blank(ws) -> int:
    """First row with nothing in column A, so demonstration rows are kept."""
    row = 2
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    return row


def clear_seed_rows(path: Path) -> int:
    """Remove every row this script added, leaving real data untouched.

    The seed rows exist so the sheet is not empty while the field data is still being
    collected. They are placeholders for a curator's judgement, not a substitute for
    it, and once the Duhok data arrives they should go — a reconstruction sitting
    alongside attested terms is worse than no reconstruction, because after a week
    nobody remembers which rows were which.

    Identified by the SEED marker in the notes column, so a row that a curator has
    edited and re-sourced is no longer a seed row and survives this.
    """
    wb = openpyxl.load_workbook(path)
    removed = 0
    for prefix in ("LEXICON", "TROPES", "EXAMPLES"):
        ws = _sheet(wb, prefix)
        # LEXICON's notes sit at column 10 with added_by at 11, so max_column is the
        # wrong cell there — the same off-by-two that let the trope importer treat a
        # boolean as a source string and import its own demonstration rows.
        notes_col = {"LEXICON": 10, "TROPES": 13, "EXAMPLES": 10}[prefix]
        for row in range(ws.max_row, 1, -1):
            note = str(ws.cell(row=row, column=notes_col).value or "")
            if note.startswith(MARKER):
                ws.delete_rows(row)
                removed += 1
    wb.save(path)
    print(f"removed {removed} seed rows from {path}")
    return 0


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    path = Path(args[0]) if args else Path("docs/lexicon_data_entry_template.xlsx")

    if "--clear" in sys.argv:
        return clear_seed_rows(path)

    wb = openpyxl.load_workbook(path)

    for prefix, rows in (("LEXICON", LEXICON), ("TROPES", TROPES), ("EXAMPLES", EXAMPLES)):
        ws = _sheet(wb, prefix)
        start = _first_blank(ws)
        for offset, row in enumerate(rows):
            if len(row) != ws.max_column and prefix != "EXAMPLES":
                raise SystemExit(
                    f"{prefix} row {offset + 1} has {len(row)} cells, "
                    f"sheet has {ws.max_column} columns"
                )
            for column, value in enumerate(row, start=1):
                ws.cell(row=start + offset, column=column, value=value)
        print(f"{prefix}: wrote {len(rows)} rows from row {start}")

    wb.save(path)
    print(f"saved {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
