"""Generate the curator data-entry workbook.

Non-technical staff should not be asked to write YAML or edit a database. This
produces an .xlsx whose columns map one-to-one onto the lexicon and trope tables, with
dropdowns on every constrained field and worked examples drawn from the real Duhok
data, so a curator fills in rows and the importer takes it from there.

Two decisions behind the layout:

* **Bilingual headers, Arabic first.** The curators read Arabic; the column names have
  to match the database for the importer. Both live in the header, Arabic on top.
* **Dropdowns rather than free text** wherever the database constrains the value.
  A typo in `target_group` produces a group that silently never matches a trope, and
  that failure is invisible — so it is prevented at entry instead.

Run:  python tools/make_lexicon_template.py [output.xlsx]
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
REQUIRED_FILL = PatternFill("solid", fgColor="C00000")
EXAMPLE_FILL = PatternFill("solid", fgColor="E7E6E6")
JUDGEMENT_FILL = PatternFill("solid", fgColor="7030A0")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TARGET_GROUPS = [
    "yazidi", "christian-iraqi", "shabak", "kakai",
    "sabian-mandaean", "turkmen-iraqi", "faili-kurd", "bahai", "kurdish",
]
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.classifiers.categories import CATEGORIES as _CATS, VISUAL_FORMS as _VIS  # noqa: E402

# Ordered by how many of the 67 survey respondents named each form, so the
# commonest sit at the top of a 14-item dropdown. A curator working in Arabic under
# time pressure should not have to scroll past rare categories to reach mockery.
CATEGORIES = [
    c.slug
    for c in sorted(_CATS, key=lambda c: -(c.reported_by or 0))
    if c.slug != "none"
]
VISUAL_FORMS = [c.slug for c in sorted(_VIS, key=lambda c: -(c.reported_by or 0))]
LANGUAGES = ["ar", "ku"]
YES_NO = ["yes", "no"]


def _style_header(ws, columns, row=1):
    for index, (arabic, english, width, kind) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=index)
        cell.value = f"{arabic}\n{english}"
        cell.fill = {
            "required": REQUIRED_FILL,
            "judgement": JUDGEMENT_FILL,
        }.get(kind, HEADER_FILL)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[row].height = 46
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _add_examples(ws, rows, start=2):
    for offset, values in enumerate(rows):
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row=start + offset, column=column, value=value)
            cell.fill = EXAMPLE_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        ws.row_dimensions[start + offset].height = 32


def _validate(ws, column_letter, options, first=2, last=500):
    validation = DataValidation(
        type="list", formula1=f'"{",".join(options)}"', allow_blank=True, showErrorMessage=True,
    )
    validation.error = "Pick a value from the list."
    validation.errorTitle = "Not a valid value"
    ws.add_data_validation(validation)
    validation.add(f"{column_letter}{first}:{column_letter}{last}")


# --------------------------------------------------------------------- sheets


def build_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("ابدأ هنا · START HERE", 0)
    ws.sheet_view.rightToLeft = True
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 118

    lines = [
        ("h", "دليل إدخال بيانات خطاب الكراهية · Hate-Speech Data Entry Guide"),
        ("", ""),
        ("h2", "ما هذا الملف · What this is"),
        ("p", "تحويل بيانات دهوك الميدانية إلى الشكل الذي تحتاجه قاعدة البيانات."),
        ("p", "Turns the Duhok field data into the shape the database needs. Each column here is a database column."),
        ("", ""),
        ("h2", "أي ورقة أستخدم؟ · Which sheet do I use?"),
        ("p", "وجدت منشوراً على فيسبوك فيه خطاب كراهية، وتعليقات مسيئة تحته. ماذا أفعل؟"),
        ("p", "Found a Facebook post with hate speech and hateful comments under it. What now?"),
        ("", ""),
        ("ex", "→ EXAMPLES sheet. The whole post and all its comments go there, one row each."),
        ("p", "ورقة EXAMPLES. المنشور وكل تعليقاته، صف لكل واحد."),
        ("", ""),
        ("p", "ثم اسأل: هل رأيت كلمة أو نمطاً جديداً غير موجود في القاموس؟"),
        ("p", "Then ask: did it contain a word or pattern the dictionary does not have yet?"),
        ("", ""),
        ("ex", "a NEW word people type      → add ONE row to LEXICON"),
        ("ex", "a NEW pattern or behaviour  → add ONE row to TROPES"),
        ("ex", "nothing new                 → EXAMPLES only. This is the normal case."),
        ("", ""),
        ("p", "لا تُدخل كل منشور في LEXICON. القاموس يحتوي على الكلمات، وليس المنشورات."),
        ("p", "Do NOT put every post into LEXICON. The dictionary holds words, not posts."),
        ("p", "You will add a few LEXICON rows a week. You will add EXAMPLES rows constantly."),
        ("", ""),
        ("h2", "الفرق بين الأوراق · The difference"),
        ("p", "LEXICON + TROPES = ما الذي يبحث عنه النظام  · WHAT the system looks for"),
        ("p", "EXAMPLES = ما هو الجواب الصحيح على محتوى حقيقي · WHAT the right answer is"),
        ("p", "Without EXAMPLES there is no way to know whether the system actually works."),
        ("", ""),
        ("h2", "القاعدة الأهم · LEXICON or TROPES?"),
        ("p", "إذا كان بإمكان شخص أن يكتب هذه العبارة حرفياً في تعليق ← ورقة LEXICON"),
        ("p", "إذا كانت وصفاً لنمط أو سلوك ← ورقة TROPES"),
        ("", ""),
        ("p", "If a person could type it word-for-word in a comment → LEXICON sheet."),
        ("p", "If it describes a pattern or behaviour → TROPES sheet."),
        ("", ""),
        ("ex", "عبدة الشيطان  ← LEXICON  (people type this)"),
        ("ex", "التمييز بالإجازات في الدوام  ← TROPES  (nobody types this sentence)"),
        ("", ""),
        ("h2", "منشور واحد = عدة صفوف · One post = several rows"),
        ("p", "منشور فيه ٥ تعليقات = ٦ صفوف: صف للمنشور، وصف لكل تعليق."),
        ("p", "A post with 5 comments = 6 rows. One for the post, one per comment."),
        ("p", "Repeat the post text on every comment row — the same comment can be hate under"),
        ("p", "one post and completely harmless under another. That context IS the judgement."),
        ("", ""),
        ("h2", "سجّل التعليقات البريئة أيضاً · Record the harmless comments too"),
        ("p", "لا تسجل المسيء فقط. التعليقات العادية هي ما يقيس الإنذارات الكاذبة."),
        ("p", "Do not record only the hateful ones. The ordinary comments are what measure"),
        ("p", "false alarms — a system judged only on hate speech looks perfect while flagging"),
        ("p", "everyone. Record the whole thread as you found it."),
        ("", ""),
        ("h2", "لا تصحح الأخطاء الإملائية · Do NOT fix spelling mistakes"),
        ("p", "إذا كتب المشارك «الشبطان» بدل «الشيطان»، اتركها كما هي وضعها في عمود variants."),
        ("p", "People really type it that way online. A 'corrected' term stops matching real posts."),
        ("", ""),
        ("h2", "الأعمدة البنفسجية تحتاج حكمك · Purple columns need YOUR judgement"),
        ("p", "لا توجد في الملفات الأصلية — أنت من يقررها. اقرأ التعليق على كل عمود."),
        ("p", "These are not in the source files. You decide them. Read each column's note."),
        ("", ""),
        ("h2", "الأعمدة الحمراء إلزامية · Red columns are required"),
        ("p", "الصف الذي ينقصه أحدها سيُرفض عند الاستيراد · A row missing one of these is rejected on import."),
        ("", ""),
        ("h2", "الصفوف الرمادية أمثلة · Grey rows are examples"),
        ("p", "مأخوذة من بياناتكم الحقيقية. احذفوها قبل التسليم أو اتركوها — المستورد يتجاهلها."),
        ("p", "Taken from your real data. Delete them before hand-off, or leave them — the importer skips them."),
        ("", ""),
        ("h2", "سؤال شائع · A question that comes up"),
        ("p", "مصطلح واحد يُستخدم ضد أكثر من فئة؟ صف واحد فقط، واكتب الفئات مفصولة بفاصلة."),
        ("p", "One term used against several groups? ONE row, with the groups comma-separated."),
        ("p", "Two rows drift apart when someone edits only one of them."),
    ]

    row = 2
    for kind, text in lines:
        cell = ws.cell(row=row, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if kind == "h":
            cell.font = Font(bold=True, size=15, color="1F3864")
            ws.row_dimensions[row].height = 26
        elif kind == "h2":
            cell.font = Font(bold=True, size=11, color="1F3864")
            ws.row_dimensions[row].height = 20
        elif kind == "ex":
            cell.font = Font(size=10, italic=True, color="7030A0")
        else:
            cell.font = Font(size=10)
        row += 1


def build_lexicon(wb: Workbook) -> None:
    ws = wb.create_sheet("LEXICON · المصطلحات")
    ws.sheet_view.rightToLeft = True

    columns = [
        ("المصطلح", "term *", 26, "required"),
        ("الفئة المستهدفة", "target_groups *", 20, "required"),
        ("اللغة", "language *", 9, "required"),
        ("النوع", "category *", 16, "judgement"),
        ("الخطورة ١-١٠", "severity_weight *", 12, "judgement"),
        ("صريح؟", "is_explicit *", 11, "judgement"),
        ("أشكال أخرى", "variants", 26, ""),
        ("لا تُعلّم عندما", "never_flag_when", 20, ""),
        ("تعبير منتظم؟", "is_regex", 10, ""),
        ("المصدر", "notes / source *", 30, "required"),
        ("أضافه", "added_by", 14, ""),
    ]
    _style_header(ws, columns)

    _add_examples(ws, [
        ["عبدة الشيطان", "yazidi, christian-iraqi", "ar", "dehumanization", 8, "yes",
         "عبده شيطان, عبدة الشبطان", "news_quotation, counter_speech", "no",
         "EXAMPLE — duhok-focus-group row 3", "—"],
        ["سبايا", "yazidi", "ar", "dehumanization", 9, "yes", "سبيا", "news_quotation", "no",
         "EXAMPLE — duhok-focus-group row 2", "—"],
        ["نسطوري", "christian-iraqi", "ar", "slur", 6, "yes",
         "نساطرة, نسطوريين, نصاطره", "academic, news_quotation", "no",
         "EXAMPLE — duhok-survey row 6", "—"],
        ["حلال قتلهم", "yazidi, christian-iraqi", "ar", "incitement", 10, "yes", "", "", "no",
         "EXAMPLE — duhok-survey row 54", "—"],
        ["شەیتان پەریس", "yazidi", "ku", "dehumanization", 8, "yes",
         "شەیتانۆک", "", "no", "EXAMPLE — duhok-survey row 51", "—"],
        ["اعوذ بالله من الشيطان الرجيم", "yazidi, christian-iraqi", "ar", "dehumanization", 8, "no",
         "اعوذ بالبه من الشيطان الرجيم", "counter_speech, news_quotation", "no",
         "EXAMPLE — duhok-survey row 7 (BENIGN unless post concerns the group)", "—"],
    ])

    _validate(ws, "C", LANGUAGES)
    _validate(ws, "D", CATEGORIES)
    _validate(ws, "F", YES_NO)
    _validate(ws, "I", YES_NO)

    notes = {
        "A1": "The exact words as typed online. Not a description.",
        "B1": "Comma-separated for a term used against several groups. See the REFERENCE sheet for valid values.",
        "D1": "slur = insult · threat = threatens harm · dehumanization = compares to animals/evil · incitement = calls for action against them",
        "E1": "1-3 offensive · 4-6 dehumanizing · 7-9 severe · 10 incitement to violence",
        "F1": "yes = abusive in ANY context. no = only abusive when the post is about that group.\n\nWHEN UNSURE, WRITE NO. A wrong 'yes' silences innocent people; a wrong 'no' just costs a reviewer 30 seconds.",
        "G1": "Real misspellings and alternate spellings, comma-separated. Do NOT correct them.",
        "H1": "Contexts where this term is NOT hate speech: news_quotation, academic, counter_speech, reclaimed",
        "J1": "Where it came from, e.g. 'duhok-focus-group row 3'. A row without this is rejected.",
    }
    for ref, text in notes.items():
        from openpyxl.comments import Comment

        ws[ref].comment = Comment(text, "Guide", width=340, height=140)


def build_tropes(wb: Workbook) -> None:
    ws = wb.create_sheet("TROPES · الأنماط")
    ws.sheet_view.rightToLeft = True

    columns = [
        ("الاسم", "name *", 26, "required"),
        ("الوصف", "description *", 40, "required"),
        ("الفئة المستهدفة", "target_groups *", 20, "required"),
        ("عبارات حرفية", "surface_forms", 30, ""),
        ("مواضيع التفعيل", "activation_topics *", 26, "judgement"),
        ("مثال مسيء", "example (HATE) *", 40, "required"),
        ("مثال بريء", "negative_example (BENIGN) *", 40, "required"),
        ("مثال دفاع", "counter_speech_example", 34, ""),
        ("الخطورة ١-١٠", "severity_weight *", 12, "judgement"),
        ("بصري؟", "is_visual", 10, ""),
        ("المصدر", "notes / source *", 26, "required"),
    ]
    _style_header(ws, columns)

    _add_examples(ws, [
        ["Devil-worship libel",
         "Invokes the devil-worship libel — a defamation historically used to justify violence, including the 2014 Sinjar genocide.",
         "yazidi, christian-iraqi",
         "اعوذ بالله من الشيطان الرجيم, عبدة الشيطان",
         "yazidi, sinjar, lalish, ايزيدي",
         "اعوذ بالله من الشيطان الرجيم  ← on a post about a Yazidi ceremony at Lalish",
         "اعوذ بالله من الشيطان الرجيم  ← on a post about a snake in a garden",
         "الإيزيديون ليسوا عبدة الشيطان، هذا افتراء",
         8, "no", "EXAMPLE — duhok-focus-group row 3"],
        ["Collective blame from one individual",
         "Treating one member's act as proof of a trait shared by the whole community.",
         "yazidi, christian-iraqi, kurdish",
         "",
         "yazidi, ايزيدي, اشوري, مسيحي",
         "كل اليزيديين عنصريين لان واحد يزيدي طلع بفيديو وحجا على المسلمين",
         "هذا الشخص تصرف بشكل عنصري",
         "",
         6, "no", "EXAMPLE — duhok-focus-group row 10"],
        ["Not an authentic people",
         "Claims the community is foreign, planted by colonial powers, or has no real history in Iraq.",
         "christian-iraqi",
         "دخلاء, ليسو شعوب اصيلة",
         "اشوري, مسيحي, نينوى",
         "الاشورين دخلاء من قبل البريطانيين وليسو شعوب اصيلة",
         "الآشوريون من أقدم شعوب المنطقة وتاريخهم موثق",
         "",
         7, "no", "EXAMPLE — duhok-survey row 3"],
        ["Grave and cross desecration",
         "Desecration of Christian graves or the cross, filmed or photographed to circulate.",
         "christian-iraqi",
         "",
         "مسيحي, اشوري, كنيسة, صليب",
         "IMAGE: person urinating on a grave marked with a cross",
         "IMAGE: news photograph documenting damage to a cemetery",
         "",
         9, "yes", "EXAMPLE — duhok-survey column BE"],
    ])

    _validate(ws, "J", YES_NO)

    from openpyxl.comments import Comment

    notes = {
        "B1": "What should the classifier look for? Write it so someone who has never seen this pattern would recognise it.",
        "D1": "Literal phrases to match, comma-separated. LEAVE EMPTY if the pattern has no fixed wording.",
        "E1": "REQUIRED. Topics in the PARENT POST that make this hateful.\n\nIf left empty the pattern NEVER fires — because without it, a phrase like 'اعوذ بالله' would flag every prayer in Iraq.",
        "F1": "A real comment where this IS hate speech. Include what post it appeared under.",
        "G1": "REQUIRED. The SAME words in a HARMLESS setting.\n\nWithout this, the system flags ordinary religious speech. This is the single most important column in the sheet.",
        "H1": "Someone quoting the insult in order to REJECT it. These must never be flagged.",
        "J1": "yes = appears in images/memes, so a text-only check cannot see it.",
    }
    for ref, text in notes.items():
        ws[ref].comment = Comment(text, "Guide", width=360, height=170)


def build_examples(wb: Workbook) -> None:
    """Real post + comment pairs, for measuring accuracy.

    Separate from LEXICON and TROPES because it answers a different question. Those
    two say *what to look for*; this says *what the right answer is* on real content,
    which is the only way to know whether the classifier works.

    One row per item to judge. A post with five hateful comments is six rows — one
    for the post, five for the comments — because each is classified separately and
    each comment is judged against the post it replies to.
    """
    ws = wb.create_sheet("EXAMPLES · أمثلة حقيقية")
    ws.sheet_view.rightToLeft = True

    columns = [
        ("الرابط", "post_url", 30, ""),
        ("نص المنشور", "parent_post_text *", 42, "required"),
        ("نوع الصف", "row_type *", 12, "required"),
        ("النص المراد تقييمه", "comment_text *", 42, "required"),
        ("الفئة المستهدفة", "target_group", 18, ""),
        ("اللغة", "dialect", 12, ""),
        ("الحكم", "label *", 12, "judgement"),
        ("النوع", "category", 16, ""),
        ("الخطورة ١-١٠", "severity", 11, "judgement"),
        ("لماذا؟ (الآلية)", "why — the mechanism *", 46, "required"),
        ("النمط", "demonstrates_trope", 24, "judgement"),
        ("المقيّم ١", "annotator_1", 12, ""),
        ("المقيّم ٢", "annotator_2", 12, ""),
        ("حالة صعبة؟", "hard_case", 11, ""),
    ]
    _style_header(ws, columns)

    post = "مراسم دينية إيزيدية في معبد لالش"
    _add_examples(ws, [
        ["https://facebook.com/example/posts/1", post, "post", post,
         "yazidi", "iraqi", "benign", "", 0,
         "EXAMPLE — the post itself is a neutral description of a ceremony",
         "", "benign", "benign", "no"],
        ["https://facebook.com/example/posts/1", post, "comment",
         "اعوذ بالله من الشيطان الرجيم", "yazidi", "iraqi", "hate", "dehumanization", 8,
         "EXAMPLE — the pious phrase is AIMED AT Yazidi religious practice, which invokes the devil-worship libel. Any phrase invoking the devil against this community works the same way",
         "Devil-worship libel", "hate", "hate", "yes"],
        ["https://facebook.com/example/posts/1", post, "comment",
         "عبدة الشيطان كلهم", "yazidi", "iraqi", "hate", "dehumanization", 8,
         "EXAMPLE — explicit slur; hateful with or without the post context",
         "Devil-worship libel", "hate", "hate", "no"],
        ["https://facebook.com/example/posts/1", post, "comment",
         "الإيزيديون ليسوا عبدة الشيطان، هذا افتراء", "yazidi", "iraqi", "benign", "counter_speech", 0,
         "EXAMPLE — quotes the libel in order to REFUTE it. Negation reverses the meaning; flagging this attacks a defender of the community",
         "Devil-worship libel", "benign", "benign", "yes"],
        ["https://facebook.com/example/posts/1", post, "comment",
         "شكرا على المعلومات", "yazidi", "iraqi", "benign", "", 0,
         "EXAMPLE — ordinary comment. The clean ones matter: they measure false alarms",
         "", "benign", "benign", "no"],
        ["https://facebook.com/other/posts/9", "شاهدوا هذا الثعبان الضخم في الحديقة", "comment",
         "اعوذ بالله من الشيطان الرجيم", "", "iraqi", "benign", "", 0,
         "EXAMPLE — THE PAIR. Same words as row 3, but the post concerns a snake, so the phrase is ordinary piety. The MECHANISM is that the trope needs the group in context",
         "Devil-worship libel", "benign", "benign", "yes"],
    ])

    _validate(ws, "C", ["post", "comment"])
    _validate(ws, "G", ["hate", "benign", "ambiguous"])
    _validate(ws, "H", CATEGORIES + ["counter_speech", "news_reporting"])
    _validate(ws, "L", ["hate", "benign", "ambiguous"])
    _validate(ws, "M", ["hate", "benign", "ambiguous"])
    _validate(ws, "N", YES_NO)
    _validate(ws, "E", TARGET_GROUPS)

    from openpyxl.comments import Comment

    notes = {
        "B1": "The POST the comment appeared under. Repeat it on every comment row from that post.\n\nThis is the most important column in the sheet: the same comment can be hate on one post and harmless on another.",
        "C1": "post = judging the post itself. comment = judging a comment under it.\n\nOne post with 5 comments = 6 rows.",
        "D1": "For a 'post' row, repeat the post text here. For a 'comment' row, the comment.",
        "G1": "hate / benign / ambiguous.\n\nambiguous is a REAL answer — use it when you genuinely cannot tell. It is not counted as a mistake.",
        "J1": "Why this is the right answer. Written for someone who was not there.\n\nFor a hard case, say what makes it hard.",
        "K1": "Two people should label the same items INDEPENDENTLY, without seeing each other's answer.\n\nThis measures whether the definition is clear enough to apply consistently. If two experts disagree often, the problem is the definition, not the model.",
        "M1": "yes = a case that only makes sense with the post context, or that a careless reader would get wrong.\n\nThese are reported separately, because overall scores can look fine while every hard case is wrong.",
    }
    for ref, text in notes.items():
        ws[ref].comment = Comment(text, "Guide", width=380, height=180)


def build_reference(wb: Workbook) -> None:
    ws = wb.create_sheet("REFERENCE · المرجع")
    ws.sheet_view.rightToLeft = True

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 56

    sections = [
        ("الفئات المستهدفة · TARGET GROUPS", ["slug", "الاسم", "ملاحظات · notes"], [
            ("yazidi", "الإيزيديون", "aliases: ايزيدي, ئێزیدی, yezidi, ezidi"),
            ("christian-iraqi", "المسيحيون / الآشوريون", "aliases: اشوري, كلداني, سرياني, نصراني"),
            ("shabak", "الشبك", ""),
            ("kakai", "الكاكائية", "also: يارسان"),
            ("sabian-mandaean", "الصابئة المندائيون", ""),
            ("turkmen-iraqi", "التركمان", ""),
            ("faili-kurd", "الكرد الفيليون", ""),
            ("bahai", "البهائيون", ""),
            ("kurdish", "الكرد", "also appears as a TARGET in the survey data"),
        ]),
        ("النوع · CATEGORY", ["value", "المعنى", "مثال · example"], [
            ("slur", "شتيمة أو لقب مهين", "نسطوري"),
            ("threat", "تهديد بالأذى", ""),
            ("dehumanization", "تشبيه بالحيوان أو الشر", "عبدة الشيطان"),
            ("incitement", "دعوة لإيذائهم", "حلال قتلهم"),
        ]),
        ("الخطورة · SEVERITY", ["1-10", "المعنى", ""], [
            ("1-3", "مسيء", "offensive"),
            ("4-6", "تجريد من الإنسانية", "dehumanizing"),
            ("7-9", "شديد", "severe"),
            ("10", "تحريض على العنف", "incitement to violence"),
        ]),
        ("لا تُعلّم عندما · NEVER_FLAG_WHEN", ["value", "المعنى", ""], [
            ("news_quotation", "اقتباس صحفي", ""),
            ("academic", "بحث أكاديمي", ""),
            ("counter_speech", "شخص يرفض الإساءة", ""),
            ("reclaimed", "استخدام داخلي للفئة نفسها", ""),
        ]),
    ]

    row = 1
    for title, headers, rows in sections:
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = HEADER_FILL
        for column in range(1, 4):
            ws.cell(row=row, column=column).fill = HEADER_FILL
        row += 1
        for column, header in enumerate(headers, start=1):
            head = ws.cell(row=row, column=column, value=header)
            head.font = Font(bold=True, size=10)
            head.border = BORDER
        row += 1
        for values in rows:
            for column, value in enumerate(values, start=1):
                body = ws.cell(row=row, column=column, value=value)
                body.border = BORDER
                body.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1


def main() -> None:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("lexicon_data_entry_template.xlsx")

    wb = Workbook()
    wb.remove(wb.active)
    build_instructions(wb)
    build_lexicon(wb)
    build_tropes(wb)
    build_examples(wb)
    build_reference(wb)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f"wrote {output}")


if __name__ == "__main__":
    main()
