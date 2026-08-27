"""Validate a filled curator workbook and convert it to pack YAML.

Closes the loop: staff fill the .xlsx, this checks it and emits
`packs/<name>/lexicon.yaml` and `tropes.yaml`, which `ankedo pack install` loads.

Validation is strict on purpose. Every rule here corresponds to a way the classifier
fails *silently* in production rather than loudly at import:

* a missing `source` produces a term nobody can defend when a report is challenged
* an unknown `target_group` produces a term that never matches its trope, with no error
* a trope with no `negative_examples` flags ordinary devout speech
* a trope with no `activation_topics` either never fires or fires on everything,
  depending on how it is read

Better to reject a row now than to discover it months later in the eval.

Run:  python tools/import_lexicon_sheet.py filled.xlsx packs/iraq-minorities
      python tools/import_lexicon_sheet.py filled.xlsx --check   # validate only

Produces lexicon.yaml, tropes.yaml and gold_eval.jsonl.
"""
from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass, field
from pathlib import Path

import yaml
from openpyxl import load_workbook

VALID_GROUPS = {
    "yazidi", "christian-iraqi", "shabak", "kakai",
    "sabian-mandaean", "turkmen-iraqi", "faili-kurd", "bahai", "kurdish",
}
# Shared with the classifier — see src/classifiers/categories.py
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.classifiers.categories import slugs as _category_slugs  # noqa: E402

VALID_CATEGORIES = set(_category_slugs())
VALID_LANGUAGES = {"ar", "ku"}
VALID_CONTEXTS = {"news_quotation", "academic", "counter_speech", "reclaimed"}

# The grey example rows shipped in the template are marked with this prefix, so a
# curator can leave them in place as a reference. Matching on the source *text* was
# too fragile — a real curated row may legitimately cite the same transcript row.
EXAMPLE_MARKER = "EXAMPLE"


@dataclass
class Result:
    lexicon: list[dict] = field(default_factory=list)
    tropes: list[dict] = field(default_factory=list)
    examples: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    skipped: int = 0

    @property
    def ok(self) -> bool:
        return not self.errors


def _text(value) -> str:
    return str(value).strip() if value is not None else ""


def _split(value) -> list[str]:
    """Split a multi-value cell. Curators use several separators interchangeably."""
    raw = _text(value)
    if not raw:
        return []
    for separator in ("،", "؛", ";", "|", "/"):
        raw = raw.replace(separator, ",")
    return [part.strip() for part in raw.split(",") if part.strip()]


def _is_example(source: str) -> bool:
    return source.upper().startswith(EXAMPLE_MARKER)


def _severity(value, where: str, result: Result) -> int | None:
    raw = _text(value)
    if not raw:
        result.errors.append(f"{where}: severity_weight is required")
        return None
    try:
        number = int(float(raw))
    except ValueError:
        result.errors.append(f"{where}: severity_weight {raw!r} is not a number")
        return None
    if not 1 <= number <= 10:
        result.errors.append(f"{where}: severity_weight {number} is outside 1-10")
        return None
    return number


def _groups(value, where: str, result: Result) -> list[str]:
    groups = _split(value)
    if not groups:
        result.errors.append(f"{where}: target_groups is required")
        return []
    unknown = [g for g in groups if g not in VALID_GROUPS]
    if unknown:
        result.errors.append(
            f"{where}: unknown target group(s) {unknown} — a term filed under an "
            f"unknown group silently never matches its trope. Valid: {sorted(VALID_GROUPS)}"
        )
    return [g for g in groups if g in VALID_GROUPS]


def read_lexicon(ws, result: Result) -> None:
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not _text(row[0]):
            continue
        source = _text(row[9]) if len(row) > 9 else ""
        if _is_example(source):
            result.skipped += 1
            continue

        where = f"LEXICON row {index}"
        term = _text(row[0])

        if not source:
            result.errors.append(
                f"{where}: notes/source is required — an unsourced term cannot be "
                "defended when a report is challenged"
            )

        language = _text(row[2]).lower()
        if language and language not in VALID_LANGUAGES:
            result.errors.append(f"{where}: language {language!r} must be ar or ku")

        category = _text(row[3]).lower()
        if category and category not in VALID_CATEGORIES:
            result.errors.append(f"{where}: category {category!r} not in {sorted(VALID_CATEGORIES)}")

        explicit_raw = _text(row[5]).lower()
        if explicit_raw not in ("yes", "no"):
            result.errors.append(
                f"{where}: is_explicit must be yes or no. When unsure write no — a "
                "wrong 'yes' flags innocent speech"
            )
        is_explicit = explicit_raw == "yes"

        contexts = [c for c in _split(row[7] if len(row) > 7 else "")]
        unknown_contexts = [c for c in contexts if c not in VALID_CONTEXTS]
        if unknown_contexts:
            result.warnings.append(
                f"{where}: unrecognised never_flag_when {unknown_contexts} — ignored"
            )

        severity = _severity(row[4], where, result)
        groups = _groups(row[1], where, result)

        if is_explicit and severity and severity >= 9 and not contexts:
            result.warnings.append(
                f"{where}: severity {severity} with no never_flag_when — a term this "
                "severe will flag journalists quoting it"
            )

        result.lexicon.append({
            "term": term,
            "target_groups": groups,
            "dialect": [language] if language else [],
            "script": ["arabic"],
            "is_explicit": is_explicit,
            "severity": severity,
            "category": category or None,
            "variants": _split(row[6] if len(row) > 6 else ""),
            "never_flag_when": [c for c in contexts if c in VALID_CONTEXTS],
            "is_regex": _text(row[8] if len(row) > 8 else "").lower() == "yes",
            "source": source,
            "added_by": _text(row[10]) if len(row) > 10 else None,
        })


def read_tropes(ws, result: Result) -> None:
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not _text(row[0]):
            continue
        # Index 12, not 10. When requires_target_group and negation_cancels were
        # inserted at K and L, notes moved to M and this was not updated — so the
        # skip test was reading a boolean, never matched "EXAMPLE", and the four
        # demonstration tropes in the template were imported as live rules.
        source = _text(row[12]) if len(row) > 12 else ""
        if _is_example(source):
            result.skipped += 1
            continue

        where = f"TROPES row {index}"
        name = _text(row[0])

        if not _text(row[1]):
            result.errors.append(f"{where}: description is required")
        if not source:
            result.errors.append(f"{where}: notes/source is required")

        groups = _groups(row[2], where, result)
        surface_forms = _split(row[3] if len(row) > 3 else "")
        topics = _split(row[4] if len(row) > 4 else "")
        positive = _text(row[5]) if len(row) > 5 else ""
        negative = _text(row[6]) if len(row) > 6 else ""

        if not positive:
            result.errors.append(f"{where}: an example of the hateful use is required")

        if not negative:
            result.errors.append(
                f"{where}: negative_example is REQUIRED. Without the same words in a "
                "harmless setting, this pattern flags ordinary speech — for a "
                "minority-protection tool that is worse than missing hate"
            )

        if not topics:
            result.errors.append(
                f"{where}: activation_topics is required. Empty means the pattern "
                "never fires, so the row would be imported and do nothing"
            )

        severity = _severity(row[8], where, result)
        is_visual = _text(row[9] if len(row) > 9 else "").lower() == "yes"

        # Read, not assumed. Both were hardcoded True below, which made the two
        # columns decorative: a curator could set them either way and nothing changed.
        #
        # Defaults when the cell is blank are the safe ones. requires_target_group
        # defaults True because a trope that fires without a target present matches
        # ordinary speech — "الخونة" is everyday political argument in Iraq.
        # negation_cancels defaults True because "they are not devil-worshippers" is a
        # denial of the libel, and flagging the person rejecting it puts a defender in
        # the evidence file.
        def _flag(index: int, default: bool) -> bool:
            raw = _text(row[index] if len(row) > index else "").lower()
            if raw in ("yes", "true", "1", "نعم"):
                return True
            if raw in ("no", "false", "0", "لا"):
                return False
            return default

        requires_target_group = _flag(10, True)
        negation_cancels = _flag(11, True)

        if is_visual and surface_forms:
            result.warnings.append(
                f"{where}: marked visual but has surface_forms — text matching will "
                "not see an image"
            )

        result.tropes.append({
            "trope_id": _slug(name),
            "target_groups": groups,
            "surface_forms": [{"text": f, "register": "unspecified"} for f in surface_forms],
            "activation": {
                "requires_target_group": requires_target_group,
                "post_topic_any": topics,
                "negation_cancels": negation_cancels,
            },
            "implicature": _text(row[1]),
            "severity": severity,
            "is_visual": is_visual,
            "positive_examples": [{"comment_text": positive}] if positive else [],
            "negative_examples": [{"comment_text": negative}] if negative else [],
            "counter_speech_examples": (
                [{"comment_text": _text(row[7])}] if len(row) > 7 and _text(row[7]) else []
            ),
            "confirmed_in_cases": [],
            "source": source,
        })


def read_examples(ws, result: Result) -> None:
    """Real post + comment pairs → gold_eval.jsonl.

    This is what measures whether the classifier works. The two things checked
    hardest are the two that make an eval set misleading rather than merely small:
    a comment recorded without its post cannot test context-dependence at all, and a
    set containing only hateful items cannot measure false alarms.
    """
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not _text(row[3]):
            continue
        why = _text(row[9]) if len(row) > 9 else ""
        if _is_example(why):
            result.skipped += 1
            continue

        where = f"EXAMPLES row {index}"
        row_type = _text(row[2]).lower() or "comment"
        text = _text(row[3])
        parent = _text(row[1])
        label = _text(row[6]).lower()

        if label not in ("hate", "benign", "ambiguous"):
            result.errors.append(f"{where}: label must be hate, benign or ambiguous")

        if row_type == "comment" and not parent:
            result.errors.append(
                f"{where}: parent_post_text is required on a comment row — without "
                "the post, this item cannot test whether context changes the verdict, "
                "which is the whole mechanism"
            )

        if not why:
            result.warnings.append(f"{where}: no reason given — a later reviewer cannot check it")

        group = _text(row[4]) or None
        if group and group not in VALID_GROUPS:
            result.errors.append(f"{where}: unknown target_group {group!r}")
            group = None

        annotators = []
        for column, name in ((11, "a1"), (12, "a2")):
            value = _text(row[column]).lower() if len(row) > column else ""
            if value in ("hate", "benign", "ambiguous"):
                annotators.append({"id": name, "label": value})

        result.examples.append({
            "id": f"sheet-{index:04d}",
            "comment_text": text,
            "parent_post_text": parent or None,
            "target_group": group,
            "dialect": _text(row[5]) or None,
            "label": label,
            "category": _text(row[7]) or None,
            "severity": _severity_optional(row[8] if len(row) > 8 else None),
            "annotators": annotators,
            "hard_case": _text(row[13]).lower() == "yes" if len(row) > 13 else False,
            # The bridge from one explained example to reusable knowledge: naming
            # the trope is what makes the explanation apply to phrasings nobody
            # has written down yet.
            "trope_id": (_text(row[10]) or None) if len(row) > 10 else None,
            "why": why,
            "source": _text(row[0]) or "curator-sheet",
        })


def _severity_optional(value) -> int | None:
    raw = _text(value)
    try:
        return int(float(raw)) if raw else None
    except ValueError:
        return None


def check_example_balance(result: Result) -> None:
    """An eval set of only hateful items cannot measure false alarms.

    A classifier that flags everything scores perfectly against it, which is the
    failure mode that silences a community rather than protecting it.
    """
    if not result.examples:
        return

    labels = [entry["label"] for entry in result.examples]
    benign = labels.count("benign")
    hate = labels.count("hate")

    if hate and not benign:
        result.warnings.append(
            f"EXAMPLES: {hate} hateful items and no benign ones. A system that flags "
            "everything would score perfectly. Record the ordinary comments too"
        )
    elif hate and benign < hate / 2:
        result.warnings.append(
            f"EXAMPLES: {hate} hateful vs {benign} benign. Aim for at least as many "
            "benign as hateful — false alarms are the harm this project can cause"
        )

    doubly = sum(1 for e in result.examples if len(e["annotators"]) >= 2)
    if result.examples and doubly < min(20, len(result.examples)):
        result.warnings.append(
            f"EXAMPLES: only {doubly} items labelled by two people. Without independent "
            "double-labelling there is no way to tell whether the definition is clear "
            "enough to apply consistently"
        )


def _slug(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in " -" else "" for ch in name.lower())
    return "-".join(cleaned.split())[:100] or "unnamed-trope"


def convert(path: Path) -> Result:
    wb = load_workbook(path, data_only=True)
    result = Result()

    lexicon_sheet = next((n for n in wb.sheetnames if n.upper().startswith("LEXICON")), None)
    trope_sheet = next((n for n in wb.sheetnames if n.upper().startswith("TROPES")), None)
    example_sheet = next((n for n in wb.sheetnames if n.upper().startswith("EXAMPLES")), None)

    if lexicon_sheet is None and trope_sheet is None and example_sheet is None:
        result.errors.append(
            "no LEXICON or TROPES sheet found — is this the right workbook?"
        )
        return result

    if lexicon_sheet:
        read_lexicon(wb[lexicon_sheet], result)
    if trope_sheet:
        read_tropes(wb[trope_sheet], result)
    if example_sheet:
        read_examples(wb[example_sheet], result)
        check_example_balance(result)

    seen: dict[str, int] = {}
    for entry in result.lexicon:
        seen[entry["term"]] = seen.get(entry["term"], 0) + 1
    for term, count in seen.items():
        if count > 1:
            result.errors.append(
                f"term {term!r} appears {count} times — use ONE row with the groups "
                "comma-separated, or two rows drift apart when only one is edited"
            )

    return result


def _write_pack_metadata(pack_dir: Path, result: Result) -> list[str]:
    """Emit the two files the installer requires and this tool never wrote.

    `ankedo pack verify` requires four files — pack.yaml, target_groups.yaml,
    lexicon.yaml and tropes.yaml — and this tool wrote two of them, then printed
    "Next: ankedo pack verify && ankedo pack install" as though the job were done. It
    was not: verification failed on the missing pair every time, so nothing the
    curator entered in the workbook could reach the agent. The last step of the only
    path from field data to a running classifier was broken.

    Existing files are never overwritten. target_groups.yaml carries aliases, adjacent
    groups and self-reference terms that no sheet in the workbook holds, so
    regenerating it from the slugs mentioned in these rows would silently delete work
    — the aliases in particular, where every missing spelling is a false negative.
    """
    written: list[str] = []

    if not (pack_dir / "pack.yaml").exists():
        (pack_dir / "pack.yaml").write_text(
            "name: imported-lexicon\n"
            "version: 0.1.0\n"
            "license: TBD                  # decide before any redistribution\n"
            "maintainer: TBD\n"
            "languages: [ar, ckb, kmr]\n"
            "region: IQ\n"
            "description: >\n"
            "  Generated by tools/import_lexicon_sheet.py from the data-entry\n"
            "  workbook. Edit name, license and maintainer before sharing this pack.\n",
            encoding="utf-8",
        )
        written.append("pack.yaml")

    if not (pack_dir / "target_groups.yaml").exists():
        slugs: set[str] = set()
        for entry in result.lexicon + result.tropes:
            slugs.update(entry.get("target_groups") or [])

        lines = [
            "# Generated from the groups referenced by the imported rows.",
            "#",
            "# `aliases` is load-bearing and is EMPTY here: a trope only fires when its",
            "# target group is matched in context, so every missing spelling,",
            "# transliteration or script variant is a silent false negative. Fill these",
            "# in before trusting per-group recall — an unaliased group will simply",
            "# never match, and will look like a group nobody attacks.",
            "",
        ]
        for slug in sorted(slugs):
            lines += [
                f"- slug: {slug}",
                f"  display_name: {{en: {slug}}}",
                "  aliases: []                   # REQUIRED before this group can match",
                "  self_reference_terms: []",
                "  adjacent_groups: []",
                "  enabled: true",
                "",
            ]
        (pack_dir / "target_groups.yaml").write_text("\n".join(lines), encoding="utf-8")
        written.append("target_groups.yaml")

    return written


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("pack_dir", type=Path, nargs="?")
    parser.add_argument("--check", action="store_true", help="validate without writing")
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"no such file: {args.workbook}")
        return 2

    result = convert(args.workbook)

    for warning in result.warnings:
        print(f"WARN  {warning}")
    for error in result.errors:
        print(f"ERROR {error}")

    print(
        f"\n{len(result.lexicon)} terms, {len(result.tropes)} tropes"
        + (f", {result.skipped} example rows skipped" if result.skipped else "")
    )

    if not result.ok:
        print(f"\n{len(result.errors)} error(s) — nothing written.")
        return 1

    if args.check or args.pack_dir is None:
        print("Valid.")
        return 0

    args.pack_dir.mkdir(parents=True, exist_ok=True)
    (args.pack_dir / "lexicon.yaml").write_text(
        yaml.safe_dump({"entries": result.lexicon}, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    (args.pack_dir / "tropes.yaml").write_text(
        yaml.safe_dump({"entries": result.tropes}, allow_unicode=True, sort_keys=False),
        encoding="utf-8",
    )
    if result.examples:
        import json

        lines = [json.dumps(entry, ensure_ascii=False) for entry in result.examples]
        (args.pack_dir / "gold_eval.jsonl").write_text(
            "\n".join(lines) + "\n", encoding="utf-8"
        )

    written = ["lexicon.yaml", "tropes.yaml"] + (["gold_eval.jsonl"] if result.examples else [])
    generated = _write_pack_metadata(args.pack_dir, result)
    written += generated

    print(f"Wrote {args.pack_dir}/: {', '.join(written)}")
    if "target_groups.yaml" in generated:
        print(
            "\nNOTE  target_groups.yaml was generated with EMPTY aliases. A trope only\n"
            "      fires when its group is matched in context, so until you fill those\n"
            "      in, those groups match nothing and will look unattacked."
        )
    print("Next:  ankedo pack verify && ankedo pack install")
    return 0


if __name__ == "__main__":
    sys.exit(main())
