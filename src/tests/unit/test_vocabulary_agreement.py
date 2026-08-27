"""The workbook teaches a vocabulary; the pack enforces a taxonomy. They must agree.

The REFERENCE sheet lists the target-group slugs a curator may write, and
`target_groups.yaml` declares the ones pack verification will accept. Nothing kept
them in step, and they had already drifted: REFERENCE offered `kurdish`, annotated
"also appears as a TARGET in the survey data", while the pack declared eight groups
and not that one. Every row a curator wrote against the sheet's own reference list was
rejected with "undeclared target group 'kurdish'".

That failure is at least loud. The same drift in the other direction is not: a group
declared in the pack but absent from REFERENCE is one no curator is told they may use,
so it silently collects nothing and reads as a community nobody attacks.

Same for the category and never_flag_when vocabularies, which the importer validates
against and the sheet advertises.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import yaml

ROOT = Path(__file__).resolve().parents[3]
WORKBOOK = ROOT / "docs" / "lexicon_data_entry_template.xlsx"
PACK_GROUPS = ROOT / "packs" / "iraq-minorities" / "target_groups.yaml"


def _reference_rows() -> list[list[str]]:
    if not WORKBOOK.exists():
        pytest.skip("workbook not present in this checkout")
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    name = next((n for n in wb.sheetnames if n.upper().startswith("REFERENCE")), None)
    assert name, "the workbook has no REFERENCE sheet"
    return [
        [("" if c is None else str(c).strip()) for c in row]
        for row in wb[name].iter_rows(values_only=True)
    ]


def _section(rows: list[list[str]], header_contains: str) -> list[str]:
    """Values from the first column of one section, between its heading and the next."""
    values: list[str] = []
    inside = False
    for row in rows:
        first = row[0] if row else ""
        if header_contains in first.upper():
            inside = True
            continue
        if inside:
            # A new section heading: an all-caps English marker in the first cell.
            if first and first.upper() == first and any(
                marker in first.upper()
                for marker in ("CATEGORY", "SEVERITY", "NEVER_FLAG", "TARGET GROUPS")
            ):
                break
            if first and first not in ("slug", "value", "1-10"):
                values.append(first)
    return values


def test_every_slug_the_sheet_offers_is_declared_in_the_pack():
    """The bug: REFERENCE offered `kurdish` and the pack did not declare it, so any
    row using the sheet's own reference list failed verification."""
    reference = set(_section(_reference_rows(), "TARGET GROUPS"))
    declared = {g["slug"] for g in yaml.safe_load(PACK_GROUPS.read_text(encoding="utf-8"))}

    undeclared = reference - declared
    assert not undeclared, (
        f"the sheet tells curators to use {sorted(undeclared)}, but pack verification "
        f"rejects them as undeclared target groups"
    )


def test_every_group_the_pack_declares_is_offered_by_the_sheet():
    """The quiet direction. A group nobody is told they may write about collects
    nothing and looks like a community nobody attacks."""
    reference = set(_section(_reference_rows(), "TARGET GROUPS"))
    declared = {g["slug"] for g in yaml.safe_load(PACK_GROUPS.read_text(encoding="utf-8"))}

    unadvertised = declared - reference
    assert not unadvertised, (
        f"the pack supports {sorted(unadvertised)} but the sheet never offers them"
    )


def test_the_categories_the_sheet_offers_are_the_ones_the_code_knows():
    from src.classifiers.categories import CATEGORIES

    reference = set(_section(_reference_rows(), "CATEGORY"))
    known = {c.slug for c in CATEGORIES}

    unknown = reference - known
    assert not unknown, f"the sheet offers categories the classifier has never heard of: {sorted(unknown)}"


def test_the_never_flag_values_the_sheet_offers_are_all_enforceable():
    """Every value a curator can write should be capable of exempting something.

    `academic` is currently the exception and is allowed here deliberately: the
    specialist has no academic category to return, so nothing can raise that signal
    yet. It is listed as a known gap rather than silently accepted — if a value is
    added to the sheet that the enforcement layer does not know at all, this fails.
    """
    from src.classifiers.exemptions import ACADEMIC, COUNTER_SPEECH, NEWS_QUOTATION, RECLAIMED

    reference = set(_section(_reference_rows(), "NEVER_FLAG"))
    enforceable = {COUNTER_SPEECH, NEWS_QUOTATION, ACADEMIC, RECLAIMED}

    unknown = reference - enforceable
    assert not unknown, (
        f"the sheet offers never_flag_when values the enforcement layer ignores "
        f"entirely: {sorted(unknown)}"
    )


def test_every_group_in_the_pack_can_actually_be_matched():
    """`aliases` is load-bearing: a trope only fires when its group is matched in
    context, so a group with no aliases matches nothing regardless of what else is
    configured for it."""
    groups = yaml.safe_load(PACK_GROUPS.read_text(encoding="utf-8"))

    unmatched = [g["slug"] for g in groups if not g.get("aliases")]
    assert not unmatched, f"groups that can never be matched in context: {unmatched}"
