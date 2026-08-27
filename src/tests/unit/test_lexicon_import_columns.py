"""The TROPES importer must read the columns the sheet actually has.

Two new booleans, `requires_target_group` and `negation_cancels`, were inserted at K
and L, pushing notes from K to M. The importer was not updated, and both faults were
silent:

1. It read the source column at index 10, which is now `requires_target_group`. That
   value never begins with "EXAMPLE", so the skip test stopped matching and the
   template's four demonstration tropes — "Devil-worship libel", "Grave and cross
   desecration" and two others, all written as teaching examples — were imported as
   live detection rules. `--check` reported eleven tropes where seven were real, and
   said "Valid."

2. It hardcoded `requires_target_group` and `negation_cancels` to True in the emitted
   pack, so the two columns did nothing whatever. A curator could set either way and
   the behaviour was identical.

Neither raised. The workbook validated, the pack looked right, and the agent would
have been detecting hate speech using demonstration data written to explain the format
— on its way to a VPS, days before deployment.
"""
from __future__ import annotations

import openpyxl
import pytest

from tools.import_lexicon_sheet import Result, read_tropes

# name, description, target_groups, surface_forms, activation_topics,
# example, negative_example, counter_speech, severity, is_visual,
# requires_target_group, negation_cancels, notes
HEADERS = [
    "name", "description", "target_groups", "surface_forms", "activation_topics",
    "example", "negative_example", "counter_speech_example", "severity_weight",
    "is_visual", "requires_target_group", "negation_cancels", "notes / source",
]

NOTES = 12
REQUIRES_TARGET_GROUP = 10
NEGATION_CANCELS = 11


def _row(over=None):
    row = {
        0: "Test trope", 1: "A description", 2: "yazidi", 3: "", 4: "ايزيدي",
        5: "the hateful example", 6: "the benign example", 7: "", 8: 5,
        9: "no", REQUIRES_TARGET_GROUP: "yes", NEGATION_CANCELS: "yes",
        NOTES: "a real source",
    }
    row.update(over or {})
    return [row.get(i, "") for i in range(13)]


def _sheet(*rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    return ws


def _read(*rows) -> Result:
    result = Result()
    read_tropes(_sheet(*rows), result)
    assert not result.errors, result.errors
    return result


def test_the_notes_column_is_read_from_the_end_not_from_index_ten():
    """The whole bug. Index 10 is a boolean now."""
    result = _read(_row({NOTES: "EXAMPLE — demonstration row"}))

    assert result.tropes == [], "a demonstration row was imported as a live rule"
    assert result.skipped == 1


def test_a_real_row_is_still_imported():
    result = _read(_row({NOTES: "duhok-survey row 12"}))

    assert len(result.tropes) == 1
    assert result.skipped == 0


def test_a_demonstration_row_is_not_rescued_by_the_boolean_columns():
    """With the old index, `requires_target_group: "yes"` was what the skip test
    examined — and "yes" does not start with EXAMPLE, so every demonstration row
    passed through regardless of what its notes said."""
    result = _read(
        _row({REQUIRES_TARGET_GROUP: "yes", NOTES: "EXAMPLE — teaching row"})
    )

    assert result.tropes == []


@pytest.mark.parametrize(
    "cell,expected",
    [("yes", True), ("no", False), ("YES", True), ("No", False), ("", True)],
)
def test_requires_target_group_comes_from_the_sheet(cell, expected):
    result = _read(_row({REQUIRES_TARGET_GROUP: cell}))

    assert result.tropes[0]["activation"]["requires_target_group"] is expected


@pytest.mark.parametrize(
    "cell,expected",
    [("yes", True), ("no", False), ("", True)],
)
def test_negation_cancels_comes_from_the_sheet(cell, expected):
    result = _read(_row({NEGATION_CANCELS: cell}))

    assert result.tropes[0]["activation"]["negation_cancels"] is expected


def test_a_blank_requires_target_group_defaults_to_requiring_one():
    """The safe default. A trope that fires with no target present matches ordinary
    speech — الخونة is everyday political argument in Iraq, and flagging it
    unconditionally floods the review queue with politics."""
    result = _read(_row({REQUIRES_TARGET_GROUP: ""}))

    assert result.tropes[0]["activation"]["requires_target_group"] is True


def test_a_blank_negation_cancels_defaults_to_cancelling():
    """"They are not devil-worshippers" is a denial of the libel. Flagging it puts
    the person rejecting the abuse into the evidence file."""
    result = _read(_row({NEGATION_CANCELS: ""}))

    assert result.tropes[0]["activation"]["negation_cancels"] is True


def test_the_two_flags_are_independent():
    """They were both hardcoded True, so a test setting only one would have passed
    against the broken importer."""
    result = _read(_row({REQUIRES_TARGET_GROUP: "no", NEGATION_CANCELS: "yes"}))

    activation = result.tropes[0]["activation"]
    assert activation["requires_target_group"] is False
    assert activation["negation_cancels"] is True


def test_the_shipped_workbook_has_the_columns_this_importer_expects():
    """Guards the pairing itself: if the sheet is regenerated with a different layout,
    this fails here rather than silently importing the wrong cells."""
    from pathlib import Path

    workbook = Path(__file__).resolve().parents[3] / "docs" / "lexicon_data_entry_template.xlsx"
    if not workbook.exists():
        pytest.skip("workbook not present in this checkout")

    wb = openpyxl.load_workbook(workbook)
    name = next(n for n in wb.sheetnames if n.upper().startswith("TROPES"))
    header = [str(c.value or "") for c in wb[name][1]]

    assert len(header) == 13, f"TROPES has {len(header)} columns, importer expects 13"
    assert "requires_target_group" in header[REQUIRES_TARGET_GROUP]
    assert "negation_cancels" in header[NEGATION_CANCELS]
    assert "notes" in header[NOTES]
