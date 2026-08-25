"""The curator workbook's dropdowns.

A curator cannot see a validation that is missing, and a slug typed into a free-text
column produces a target group that silently never matches. These assert the
constraints exist in the generated .xlsx rather than in the builder's intent.
"""
from __future__ import annotations

import openpyxl
import pytest

from tools.make_lexicon_template import (
    CATEGORIES,
    HATE_CATEGORIES,
    LANGUAGES,
    NEVER_FLAG_WHEN,
    TARGET_GROUPS,
    build_workbook,
)


@pytest.fixture(scope="module")
def wb(tmp_path_factory):
    path = tmp_path_factory.mktemp("wb") / "template.xlsx"
    build_workbook(path)
    return openpyxl.load_workbook(path)


def _sheet(wb, prefix):
    return wb[next(n for n in wb.sheetnames if n.startswith(prefix))]


def _validation(ws, column):
    found = [
        dv for dv in ws.data_validations.dataValidation
        if str(dv.sqref).startswith(f"{column}2:")
    ]
    assert found, f"no validation on column {column} of {ws.title}"
    return found[0]


def _options(dv):
    return (dv.formula1 or "").strip('"').split(",")


# ── The multi-value columns ──────────────────────────────────────────────────


@pytest.mark.parametrize("prefix,column", [("LEXICON", "B"), ("TROPES", "C")])
def test_target_groups_is_constrained_on_both_sheets(wb, prefix, column):
    dv = _validation(_sheet(wb, prefix), column)
    assert _options(dv) == TARGET_GROUPS


@pytest.mark.parametrize("prefix,column", [("LEXICON", "B"), ("TROPES", "C")])
def test_target_groups_warns_rather_than_blocks(wb, prefix, column):
    """It holds several comma-separated groups; a hard list would reject them all."""
    assert _validation(_sheet(wb, prefix), column).errorStyle == "warning"


def test_never_flag_when_is_constrained_but_not_blocking(wb):
    dv = _validation(_sheet(wb, "LEXICON"), "H")
    assert _options(dv) == NEVER_FLAG_WHEN
    assert dv.errorStyle == "warning"  # "news_quotation, counter_speech" is valid


def test_multi_value_columns_explain_themselves_on_entry(wb):
    dv = _validation(_sheet(wb, "LEXICON"), "B")
    assert dv.showInputMessage
    assert "comma" in (dv.prompt or "").lower()


# ── Categories ───────────────────────────────────────────────────────────────


def test_lexicon_categories_exclude_the_non_hate_labels(wb):
    """A term in a hate dictionary cannot be categorised 'news_reporting'."""
    options = _options(_validation(_sheet(wb, "LEXICON"), "D"))
    assert "counter_speech" not in options
    assert "news_reporting" not in options
    assert options == HATE_CATEGORIES


def test_examples_keeps_the_non_hate_labels_without_duplicating_them(wb):
    options = _options(_validation(_sheet(wb, "EXAMPLES"), "H"))
    assert "counter_speech" in options
    assert len(options) == len(set(options)), "a category appears twice in the dropdown"
    assert options == CATEGORIES


def test_reference_documents_every_category_the_dropdown_offers(wb):
    """Nine of thirteen used to be undocumented, mockery — the commonest — among them."""
    ws = _sheet(wb, "REFERENCE")
    documented = {ws.cell(r, 1).value for r in range(1, ws.max_row + 1)}
    missing = [slug for slug in CATEGORIES if slug not in documented]
    assert not missing, f"undocumented in REFERENCE: {missing}"


def test_every_documented_category_carries_its_arabic_gloss(wb):
    ws = _sheet(wb, "REFERENCE")
    glosses = {
        ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)
    }
    for slug in CATEGORIES:
        assert (glosses.get(slug) or "").strip(), f"{slug} has no Arabic meaning"


# ── Trope gating columns ─────────────────────────────────────────────────────


def test_tropes_can_express_the_two_gates_that_default_to_true(wb):
    ws = _sheet(wb, "TROPES")
    headers = [(ws.cell(1, i).value or "").split("\n")[-1] for i in range(1, ws.max_column + 1)]
    assert "requires_target_group" in headers
    assert "negation_cancels" in headers


@pytest.mark.parametrize("column", ["K", "L"])
def test_the_gate_columns_are_yes_no(wb, column):
    assert _options(_validation(_sheet(wb, "TROPES"), column)) == ["yes", "no"]


# ── Language ─────────────────────────────────────────────────────────────────


def test_language_covers_where_the_speech_actually_arrives(wb):
    """The column is the language the TERM is written in, not the platform's locales.

    Hate speech against these communities arrives in Arabic and Kurdish; a slur
    aimed at Assyrians is written in Arabic, so a Syriac option would be a column
    nobody can fill correctly.
    """
    assert _options(_validation(_sheet(wb, "LEXICON"), "C")) == LANGUAGES == ["ar", "ku"]
